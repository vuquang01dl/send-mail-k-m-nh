import pandas as pd
import os
import datetime
import win32com.client as win32
from openpyxl import load_workbook
from xlsx2html import xlsx2html

def clear_data_excel(ws, start_row: int, end_row: int):
    """Xóa dữ liệu trong khoảng dòng chỉ định."""
    for row in range(start_row, end_row):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col, value="")

def df_to_excel(df, ws, start_row):
    """Ghi dữ liệu từ DataFrame vào sheet Excel."""
    for r_idx, row in enumerate(df.values, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, str) and '%' in value:
                try:
                    value = float(value.replace('%', '').strip()) / 100
                except ValueError:
                    pass
            else:
                try:
                    value = float(value)
                except ValueError:
                    pass
            ws.cell(row=r_idx, column=c_idx + 1, value=value)

def read_emails_from_file(filename):
    """Đọc danh sách email từ file .txt."""
    with open(filename, "r") as file:
        return [line.strip() for line in file.readlines() if line.strip()]

if __name__ == "__main__":
    # Lấy thời gian hiện tại
    timeStart = datetime.datetime.now()
    time_now = datetime.datetime.now()

    # Đường dẫn
    base_dir = r"C:\Users\V5030587\OneDrive - quantacn.com\Desktop\excel_report"
    file_path = os.path.join(base_dir, "2025050813.xlsx")
    html_path = os.path.join(base_dir, "html_path", "table_Sheet1.html")

    # Đọc dữ liệu từ file Excel
    wb = load_workbook(file_path)

    # Kiểm tra danh sách sheet trong file
    print(wb.sheetnames)

    # Lấy sheet đầu tiên
    if wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        print("File Excel không có sheet nào.")
        exit()

    # Ghi dữ liệu vào Excel (nếu cần cập nhật lại nội dung)
    start_row = 5
    end_row = 11
    clear_data_excel(ws, start_row, end_row)
    df_to_excel(df, ws, start_row)
    wb.save(file_path)

    # Chuyển sang HTML
    xlsx2html(file_path, html_path, locale='en', sheet=0)

    # Soạn nội dung email
    body = ('<p style="font-size:10pt;font-family:Calibri,sans-serif;margin:0;"><b><span style="font-size:16pt;font-family:Times;" lang="en-US">Dear all:</span></b></p>'
            '<p style="font-size:10pt;font-family:Calibri,sans-serif;margin:0;">&nbsp;</p>'
            '<p style="font-size:10pt;font-family:Calibri,sans-serif;margin:0;"> <span style="font-size:14pt;font-family:Times;">Update FATP CTO output status. Cut off time: <span style="color:black;">' + time_now.strftime("%Y-%m-%d") + '</span> <b style="color:red;font-size:16pt;">' + time_now.strftime("%H:%M") + '</b> </span> <span style="font-size:14pt;font-family:Times;color:black;">Thanks!</span> </p>'
            '<p style="font-size:10pt;font-family:Calibri,sans-serif;margin:0;">&nbsp;</p>'
            '<p><b><span style="font-size: 18pt; font-family: Arial, sans-serif; color: blue;">(Sheet1):</span></b></p>')

    with open(html_path, 'r', encoding='utf-8') as file:
        body += re.sub('cellpadding="0"', 'cellpadding="10"', file.read())

    body += ('<p>&nbsp;</p>'
             f"<div><p>Time start: {timeStart.strftime('%Y-%m-%d %H:%M:%S')}</p>"
             f"<p>Time finish: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>"
             "<p style='color:red'><b>MIS负责RPA开发，使用单位需依附件确认RPA执行结果，并处理异常</b></p>"
             "<p style='color:red'><b>MIS chịu trách nhiệm phát triển RPA, đơn vị người dùng phải xác nhận kết quả thực hiện RPA theo tệp đính kèm và xử lý bất thường.</b></p></div>")

    # Đọc danh sách người nhận và cc từ file txt
    to_list = read_emails_from_file(os.path.join(base_dir, "recipients.txt"))
    cc_list = read_emails_from_file(os.path.join(base_dir, "cc.txt"))

    # Gửi email bằng Outlook
    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)
    mail.To = "; ".join(to_list)
    mail.CC = "; ".join(cc_list)
    mail.Subject = f"QMH FATP OCT Performance_{time_now.strftime('%Y-%m-%d')}"
    mail.HTMLBody = body

    # Đính kèm file Excel
    mail.Attachments.Add(Source=file_path)

    mail.Send()

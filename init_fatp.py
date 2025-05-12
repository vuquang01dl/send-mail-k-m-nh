import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from xlsx2html import xlsx2html

def clear_data_excel(ws, start_row: int, end_row: int):
    for row in range(start_row, end_row):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col, value="")

def df_to_excel(df, ws, start_row):
    for r_idx, row in enumerate(df.values, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, str) and '%' in value:
                try:
                    value = float(value.replace('%', '').strip()) / 100
                except ValueError:
                    pass
            else:
                try:
                    value = float(value)
                except ValueError:
                    pass
            ws.cell(row=r_idx, column=c_idx + 1, value=value)

def copy_cell_styles(source_ws, target_ws, start_row, end_row):
    for row in range(start_row, end_row):
        for col in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=row, column=col)
            target_cell = target_ws.cell(row=row, column=col)

            # Sao chép màu nền
            if source_cell.fill:
                target_cell.fill = source_cell.fill

            # Sao chép phông chữ
            if source_cell.font:
                target_cell.font = source_cell.font

            # Sao chép đường viền
            if source_cell.border:
                target_cell.border = source_cell.border

            # Sao chép căn lề
            if source_cell.alignment:
                target_cell.alignment = source_cell.alignment

            # Sao chép định dạng số
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

if __name__ == "__main__":
    timeStart = str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    time_now = datetime.datetime.now()

    # Path đến file Excel và HTML của bạn
    file_path = r"C:\Users\V5030587\OneDrive - quantacn.com\Desktop\excel_report\456.xlsx"
    html_path_J614 = r"C:\Users\V5030587\OneDrive - quantacn.com\Desktop\excel_report\table_J614.html"

    # Đọc file Excel
    wb = load_workbook(file_path)
    ws_J614 = wb["Sheet1"]

    start_row = 5
    end_row = 11
    clear_data_excel(ws_J614, start_row, end_row)

    # Chuyển DataFrame sang Excel
    # Giả sử bạn đã có DataFrame từ dữ liệu Excel sẵn có

    # Sao chép kiểu dữ liệu từ sheet gốc sang sheet mới
    copy_cell_styles(ws_J614, ws_J614, start_row, end_row)

    # Lưu lại file Excel
    wb.save(file_path)

    # Chuyển đổi Excel sang HTML với việc giữ nguyên màu nền
    xlsx2html(file_path, html_path_J614, locale='en', sheet=0)
